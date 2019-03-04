from random import randint
from math import floor
import openpyxl
from docRun import createDoc

def qList(difficulty,noq):
    que=floor((difficulty/100*noq))
    cost=(que*5)+(noq-que)-1
    q=[]
    for i in range(noq):
        r=randint(1,5)
        r=r%cost
        if r==0:
            q.append(1)
        else:
            q.append(r)
        cost-=r
    return q

def excel(diff,sheetNo,location):
    wb = openpyxl.load_workbook(location, read_only=True)
    sheets = wb.sheetnames

    ws = wb[sheets[sheetNo]]
    a = []
    for row in ws.iter_rows(min_row=1, max_col=5, max_row=1, values_only=True):
        a.append(row)
    a = list(a[0])
    b = []
    for i in range(len(a)):
        s = 2
        for j in range(i):
            s += a[j]
        b.append((a[i], s))   # Tuple is generated

    ques = []

    if sheetNo == 0:
        for i in range(len(diff)):
            r=randint(1,b[diff[i]-1][0])
            r+= b[diff[i]-1][1]
            for row in ws.iter_rows(min_row=r, min_col=2, max_col=6, max_row=r, values_only=True):
                ques.append(row)
        return ques
    elif len(diff)==2:
        for i in range(len(diff[0])):
            r = randint(1, b[diff[0][i] - 1][0])
            r += b[diff[0][i] - 1][1]
            cell = "B{}".format(r)

            r = randint(1, b[diff[1][i] - 1][0])
            r += b[diff[1][i] - 1][1]
            cell2 = "B{}".format(r)
            ques.append((ws[cell].value,ws[cell2].value))
        return ques
    else:
        for i in range(len(diff)):
            r = randint(1, b[diff[i] - 1][0])
            r += b[diff[i] - 1][1]
            cell = "B{}".format(r)
            ques.append(ws[cell].value)
        return ques


def questions(values):
    wb = openpyxl.load_workbook(values[0])

    if values[2] == "University Exam":

        mcq=qList((int)(values[3]),20)
        marks4=qList((int)(values[3]),7)
        marks12=[
                    qList((int)(values[3]),5),
                    qList((int)(values[3]),5)
                ]
        list=[values[2]]
        list.append(excel(mcq, 0, values[0]))
        list.append(excel(marks4, 3, values[0]))
        list.append(excel(marks12, 7, values[0]))

        if createDoc(list,values[1]) is True:
            return True
        else:
            return False

    else:
        marks4=qList((int)(values[3]),7)
        marks10=[
                    qList((int)(values[3]),3),
                    qList((int)(values[3]),3)
                ]
        list = [values[2]]
        list.append(excel(marks4, 3, values[0]))
        list.append(excel(marks10, 6, values[0]))
        if createDoc(list, values[1]) is True:
            return True
        else:
            return False
